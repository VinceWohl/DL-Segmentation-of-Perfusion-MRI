#!/usr/bin/env python3
"""
Data Completeness Checker

This script checks the completeness of the neuroimaging data structure
and generates a CSV report of missing files and folders.

Expected structure:
- DATA_HC: subjects p001-p015 with First_visit, Second_visit, Third_visit
- DATA_patients: subjects p016-p023 with First_visit, Second_visit

Author: Generated for data quality assurance
"""

import os
import csv
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class DataCompletenessChecker:
    def __init__(self, data_root_path):
        self.data_root = Path(data_root_path)
        self.results = []
        
        # Define expected structure
        self.hc_subjects = [f"sub-p{i:03d}" for i in range(1, 16)]  # p001-p015
        self.patient_subjects = [f"sub-p{i:03d}" for i in range(16, 24)]  # p016-p023
        self.hc_visits = ["First_visit", "Second_visit", "Third_visit"]
        self.patient_visits = ["First_visit", "Second_visit"]
        
        # Expected files for each subject
        self.expected_files = [
            "FLAIR/anon_{subject}_FLAIR.nii",
            "T1w/anon_{subject}_T1w.nii",
            "task-AIR/ASL/ssLICA/T1w_coreg/anon_r{subject}_T1w.nii",
            "task-AIR/ASL/ssRICA/T1w_coreg/anon_r{subject}_T1w.nii",
            "task-AIR/ASL/ssLICA/CBF_nativeSpace/CBF_3_BRmsk_CSF.nii",
            "task-AIR/ASL/ssRICA/CBF_nativeSpace/CBF_3_BRmsk_CSF.nii",
            "task-AIR/ASL/ssLICA/PerfTerrMask/mask_LICA_manual_Corrected.nii",
            "task-AIR/ASL/ssRICA/PerfTerrMask/mask_RICA_manual_Corrected.nii"
        ]
    
    def check_file_exists(self, file_path):
        """Check if a file exists"""
        return file_path.exists()
    
    def check_subject_completeness(self, group, subject, visits):
        """Check completeness for a single subject across all visits"""
        for visit in visits:
            visit_path = self.data_root / group / visit / "output" / subject
            
            # Initialize result row
            result_row = {
                'Group': group,
                'Subject': subject,
                'Visit': visit,
                'T1w': '-',
                'FLAIR': '-',
                'T1w_coreg_L': '-',
                'T1w_coreg_R': '-',
                'FLAIR_coreg_L': '-',
                'FLAIR_coreg_R': '-',
                'CBF_L': '-',
                'CBF_R': '-',
                'PerfTerrMask_L': '-',
                'PerfTerrMask_R': '-'
            }
            
            if visit_path.exists():
                # Check T1w
                t1w_path = visit_path / f"T1w/anon_{subject}_T1w.nii"
                result_row['T1w'] = '+' if self.check_file_exists(t1w_path) else '-'
                
                # Check FLAIR
                flair_path = visit_path / f"FLAIR/anon_{subject}_FLAIR.nii"
                result_row['FLAIR'] = '+' if self.check_file_exists(flair_path) else '-'
                
                # Check T1w_coreg files
                t1w_coreg_l_path = visit_path / f"task-AIR/ASL/ssLICA/T1w_coreg/anon_r{subject}_T1w.nii"
                result_row['T1w_coreg_L'] = '+' if self.check_file_exists(t1w_coreg_l_path) else '-'
                
                t1w_coreg_r_path = visit_path / f"task-AIR/ASL/ssRICA/T1w_coreg/anon_r{subject}_T1w.nii"
                result_row['T1w_coreg_R'] = '+' if self.check_file_exists(t1w_coreg_r_path) else '-'
                
                # Check FLAIR_coreg files
                flair_coreg_l_path = visit_path / f"task-AIR/ASL/ssLICA/FLAIR_coreg/anon_r{subject}_FLAIR.nii"
                result_row['FLAIR_coreg_L'] = '+' if self.check_file_exists(flair_coreg_l_path) else '-'
                
                flair_coreg_r_path = visit_path / f"task-AIR/ASL/ssRICA/FLAIR_coreg/anon_r{subject}_FLAIR.nii"
                result_row['FLAIR_coreg_R'] = '+' if self.check_file_exists(flair_coreg_r_path) else '-'
                
                # Check CBF files
                cbf_l_path = visit_path / "task-AIR/ASL/ssLICA/CBF_nativeSpace/CBF_3_BRmsk_CSF.nii"
                result_row['CBF_L'] = '+' if self.check_file_exists(cbf_l_path) else '-'
                
                cbf_r_path = visit_path / "task-AIR/ASL/ssRICA/CBF_nativeSpace/CBF_3_BRmsk_CSF.nii"
                result_row['CBF_R'] = '+' if self.check_file_exists(cbf_r_path) else '-'
                
                # Check PerfTerrMask files
                mask_l_path = visit_path / "task-AIR/ASL/ssLICA/PerfTerrMask/mask_LICA_manual_Corrected.nii"
                result_row['PerfTerrMask_L'] = '+' if self.check_file_exists(mask_l_path) else '-'
                
                mask_r_path = visit_path / "task-AIR/ASL/ssRICA/PerfTerrMask/mask_RICA_manual_Corrected.nii"
                result_row['PerfTerrMask_R'] = '+' if self.check_file_exists(mask_r_path) else '-'
            
            self.results.append(result_row)
    
    def run_completeness_check(self):
        """Run the complete data structure check"""
        print(f"Checking data completeness in: {self.data_root}")
        print("=" * 60)
        
        if not self.data_root.exists():
            print(f"ERROR: Data root directory does not exist: {self.data_root}")
            return False
        
        # Check HC subjects
        print("Checking DATA_HC subjects...")
        for subject in self.hc_subjects:
            self.check_subject_completeness("DATA_HC", subject, self.hc_visits)
            print(f"  {subject}: OK")
        
        # Check Patient subjects
        print("\nChecking DATA_patients subjects...")
        for subject in self.patient_subjects:
            self.check_subject_completeness("DATA_patients", subject, self.patient_visits)
            print(f"  {subject}: OK")
        
        # Calculate summary statistics
        total_files = 0
        missing_files = 0
        
        for result in self.results:
            for file_type in ['T1w', 'FLAIR', 'T1w_coreg_L', 'T1w_coreg_R', 'FLAIR_coreg_L', 'FLAIR_coreg_R', 'CBF_L', 'CBF_R', 'PerfTerrMask_L', 'PerfTerrMask_R']:
                total_files += 1
                if result[file_type] == '-':
                    missing_files += 1
        
        existing_files = total_files - missing_files
        completeness_pct = (existing_files / total_files * 100) if total_files > 0 else 0
        
        print("\n" + "=" * 60)
        print("SUMMARY:")
        print(f"Total entries processed: {len(self.results)}")
        print(f"Total files expected: {total_files}")
        print(f"Total files found: {existing_files}")
        print(f"Total files missing: {missing_files}")
        print(f"Overall completeness: {completeness_pct:.1f}%")
        
        return True
    
    def save_excel_report_only(self, output_path=None):
        """Save detailed results to Excel file with colored formatting"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = Path(__file__).parent / f"data_completeness_report_{timestamp}.xlsx"
        
        if not EXCEL_AVAILABLE:
            print("ERROR: openpyxl package is required. Please install with: pip install openpyxl")
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Completeness"
        
        # Define colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light gray
        header_font = Font(bold=True)
        
        # Write headers
        headers = ['Group', 'Subject', 'Visit', 'T1w', 'FLAIR', 'T1w_coreg_L', 'T1w_coreg_R', 'FLAIR_coreg_L', 'FLAIR_coreg_R', 'CBF_L', 'CBF_R', 'PerfTerrMask_L', 'PerfTerrMask_R']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data with formatting
        file_columns = ['T1w', 'FLAIR', 'T1w_coreg_L', 'T1w_coreg_R', 'FLAIR_coreg_L', 'FLAIR_coreg_R', 'CBF_L', 'CBF_R', 'PerfTerrMask_L', 'PerfTerrMask_R']
        
        for row_idx, result in enumerate(self.results, 2):  # Start from row 2
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=result[header])
                
                # Apply color formatting for file status columns
                if header in file_columns:
                    if result[header] == '+':
                        cell.fill = green_fill
                    elif result[header] == '-':
                        cell.fill = red_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)  # Cap width at 20
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        print(f"\nExcel report with colored formatting saved to: {output_path}")
        return output_path


def main():
    """Main function to run the completeness check"""
    # Default data path
    default_data_path = r"C:\Users\Vincent Wohlfarth\Data\anon_Data_250808"
    
    # Allow command line argument for data path
    if len(sys.argv) > 1:
        data_path = sys.argv[1]
    else:
        data_path = default_data_path
    
    print("Data Completeness Checker")
    print("=" * 60)
    
    # Initialize checker
    checker = DataCompletenessChecker(data_path)
    
    # Run completeness check
    success = checker.run_completeness_check()
    
    if success:
        # Save Excel report only
        checker.save_excel_report_only()
        print("\nCompleteness check completed successfully!")
    else:
        print("Completeness check failed!")
        return 1
    
    return 0


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)